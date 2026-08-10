# -*- coding: utf-8 -*-
"""테차상품가격리스트.xlsx -> techa-products.json 생성기.

147개 SKU를 제품 라인으로 묶고, 라인 단위로 다축 태그를 붙인다.
색상/사이즈/전원방식 같은 변형은 라인 아래 variants로 매단다.
"""
import openpyxl, json, io, sys, re, os
from collections import OrderedDict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# 이 스크립트(scripts/) 기준 상대경로 — 어느 컴퓨터에서 어느 위치로 clone해도 그대로 동작
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(REPO_ROOT, 'docs', '테차상품가격리스트.xlsx')
TAGS_SRC = os.path.join(REPO_ROOT, 'docs', 'gift-finder-tags.xlsx')
OUT = os.path.join(REPO_ROOT, 'assets', 'data', 'techa-products.json')
IMG_DIR = os.path.join(REPO_ROOT, 'assets', 'products')

# 상품 사진은 assets/products/{라인 id}.{webp|jpg|png} 규칙으로 파일명만 맞춰 넣으면
# 아래 scan_images()가 빌드 때 자동으로 찾아 image 필드에 넣는다.
# xlsx에 경로를 적는 칸을 따로 두지 않는 이유: 파일과 목록이 어긋날 여지를 아예 없애기 위함.
IMG_EXTS = ('.webp', '.jpg', '.jpeg', '.png')  # 앞쪽 우선

# ---------------------------------------------------------------------------
# 제품 라인 정의 — 매칭 구조만 코드에 둔다.
#   match: SKU 이름이 이 조건을 만족하면 해당 라인에 속함 (순서대로 평가, 먼저 맞는 라인 승)
#   all=반드시 포함할 토큰 / none=포함되면 안 되는 토큰
# 추천 태그(recipient/occasion/giftType/situation/reason)는 코드가 아니라
# docs/gift-finder-tags.xlsx에서 읽는다 — 태그를 고칠 땐 그 xlsx만 고치면 된다.
# ---------------------------------------------------------------------------
LINES = [
    dict(id="rose-hydrangea-bouquet", name="장미수국 꽃다발", cat="꽃다발",
         match=dict(all=["장미수국", "꽃다발"])),
    dict(id="hydrangea-bouquet", name="수국 꽃다발", cat="꽃다발",
         match=dict(all=["수국", "꽃다발"], none=["장미수국"])),
    dict(id="superior-rose", name="수페리어 장미", cat="꽃다발",
         match=dict(all=["수페리어"])),
    dict(id="rose-soap-bouquet", name="장미 비누꽃다발", cat="꽃다발",
         match=dict(all=["장미", "비누꽃다발"])),
    dict(id="pinkpeach-bouquet", name="핑크피치 꽃다발", cat="꽃다발",
         match=dict(all=["핑크피치"])),
    dict(id="sunflower-bouquet", name="해바라기 꽃다발", cat="꽃다발",
         match=dict(all=["해바라기", "꽃다발"])),
    dict(id="carnation-rose-bouquet", name="카네이션장미 꽃다발", cat="꽃다발",
         match=dict(all=["카네이션장미", "꽃다발"])),
    dict(id="carnation-money-bouquet", name="카네이션 돈꽃다발", cat="돈꽃다발",
         match=dict(all=["카네이션", "돈꽃다발"])),
    dict(id="rose-money-bouquet", name="장미 돈꽃다발", cat="돈꽃다발",
         match=dict(all=["장미", "돈꽃다발"])),
    dict(id="single-stem-soap-box", name="비누꽃 한송이·두송이 박스", cat="소품",
         match=dict(any_of=["장미한송이", "장미두송이", "카네이션한송이", "카네이션두송이"])),
    dict(id="kinderjoy-doll-bouquet", name="킨더조이 레보니 인형꽃다발", cat="인형꽃다발",
         match=dict(all=["킨더조이"])),
    dict(id="character-doll-bouquet", name="캐릭터 인형꽃다발", cat="인형꽃다발",
         match=dict(any_of=["마이멜로디", "시나모롤"])),
    dict(id="rebony-trophy-doll-bouquet", name="레보니 인형꽃다발 (트로피포트)", cat="인형꽃다발",
         match=dict(any_of=["깡총커플", "앙증맞은 레보니"])),
    dict(id="glassdome-moodlamp", name="유리돔 무드등", cat="무드등",
         match=dict(any_of=["안개꽃핑크 무드등", "안개꽃파랑 무드등", "스타티스 무드등",
                            "핑크장미 무드등", "레드장미 무드등", "파랑장미 무드등"])),
    dict(id="classic-money-box", name="클래식투명 용돈박스", cat="용돈박스",
         match=dict(all=["클래식투명용돈박스"])),
    dict(id="premium-window-money-box", name="프리미엄창문 용돈박스", cat="용돈박스",
         match=dict(all=["프리미엄창문"])),
    dict(id="sunflower-frame", name="해바라기 액자", cat="액자",
         match=dict(all=["해바라기 액자"])),
    dict(id="cup-moodlamp-single", name="프리저브드 꽃 무드등 한송이", cat="무드등",
         match=dict(all=["프리저브드 꽃 무드등 한송이"])),
    dict(id="ionantha-moodlamp", name="이오난사 스칸디아모스 무드등", cat="무드등",
         match=dict(all=["이오난사"])),
    dict(id="centerpiece", name="센터피스", cat="센터피스",
         match=dict(all=["센터피스"])),
    dict(id="topiary-moodlamp", name="토피어리 수국 무드등", cat="무드등",
         match=dict(all=["토피어리"])),
    dict(id="money-cake", name="용돈케이크 머니박스", cat="용돈박스",
         match=dict(all=["용돈케이크"])),
    dict(id="hydrangea-moodlamp", name="수국꽃무드등", cat="무드등",
         match=dict(all=["수국꽃무드등"])),
    dict(id="flower-postcard", name="꽃엽서카드", cat="소품",
         match=dict(all=["꽃엽서카드"])),
]

# 2026-08-04부터 스마트스토어 링크는 xlsx의 "사이트링크" 열(E)에서 직접 읽는다 —
# 라인 안 아무 SKU 행에나 하나 채워두면 그 라인의 url이 된다(대표 1개 행이면 충분,
# 색상·사이즈는 같은 페이지의 옵션이므로 SKU마다 다 채울 필요 없음).
# URL_MAP은 xlsx에 없는 라인(예: 플라워클래스처럼 EXTRA_LINES에만 있는 것)을 위한
# 수동 폴백이다 — xlsx 쪽에 값이 있으면 그게 항상 우선한다.
URL_MAP = {
    "flower-class": "https://smartstore.naver.com/itecha/products/10537190482",
}

# xlsx에 없지만 실제 운영 중인 상품 (script-guide.md 기준) — 가격 정보만 여기 둔다.
EXTRA_LINES = [
    dict(id="flower-class", name="플라워클래스 (원데이)", cat="체험",
         variants=[
             dict(sku="원데이 클래스 70분 (1인)", price=50000, option="1인"),
             dict(sku="원데이 클래스 70분 (4인 이상, 1인당)", price=45000, option="4인 이상"),
         ]),
]

# ---------------------------------------------------------------------------
COLORS = ['블렉앤화이트', '블랙앤화이트', '트렌치바이올렛', '오리지날레드', '피치코랄',
          '라벤더퍼플', '소프트핑크', '스카이블루', '펄화이트', '황금옐로', '핑크핑크',
          '베이지&블랙', '화이트&핑크', '맑은배경', '황금배경', '블루자갈', '핑크자갈',
          '화이트자갈', '딥레드', '파랑', '보라', '레드', '핑크', '화이트', '블루',
          '노랑', '피치', '샤넬']
SIZES = ['중대형', '대형', '미니', '라지', '스몰', 'A4', 'A3', '디럭스',
         '5송이', '7송이', '10송이', '15송이', '20송이', '23송이', '25송이',
         '한송이', '두송이']
# 단 수는 사이즈와 별개 축(용돈케이크는 1단/2단이 가격을 가르는 실제 변수)
TIERS = ['1단', '2단']
OPTS = ['건전지방식', 'USB방식', '돈티슈', '머니캡']


def extract(name, pool):
    return next((t for t in pool if t in name), None)


def matches(name, m):
    if 'all' in m and not all(t in name for t in m['all']):
        return False
    if 'any_of' in m and not any(t in name for t in m['any_of']):
        return False
    if 'none' in m and any(t in name for t in m['none']):
        return False
    return True


def price_band(p):
    # 구간 그대로 자른다(반올림 금지). "2만원대"는 20,000~29,999원을 뜻하므로
    # 19,000원을 2만원대로, 49,000원을 5만원 이상으로 올려보내면 라벨이 사실과 달라진다.
    if p < 20000: return "1만원대"
    if p < 30000: return "2만원대"
    if p < 40000: return "3만원대"
    if p < 50000: return "4만원대"
    return "5만원 이상"


URL_RE = re.compile(r'^https?://')


def split_tags(s):
    return [t.strip() for t in str(s or '').split(',') if t.strip()]


def scan_images():
    """assets/products/ 를 훑어 {라인 id: 웹 경로} 맵을 만든다.

    파일이 없으면 빈 맵 — 사진이 하나도 없어도 빌드는 그대로 성공한다.
    """
    found = {}
    if not os.path.isdir(IMG_DIR):
        return found
    for fname in os.listdir(IMG_DIR):
        stem, ext = os.path.splitext(fname)
        if ext.lower() not in IMG_EXTS:
            continue
        # 같은 id로 여러 확장자가 있으면 IMG_EXTS 앞쪽(webp)을 우선
        prev = found.get(stem)
        if prev and IMG_EXTS.index(os.path.splitext(prev)[1].lower()) <= IMG_EXTS.index(ext.lower()):
            continue
        found[stem] = '/assets/products/' + fname
    return found


def load_tags():
    """docs/gift-finder-tags.xlsx -> {라인ID: {recipient, occasion, giftType, situation, reason, note}}"""
    wb = openpyxl.load_workbook(TAGS_SRC, data_only=True)
    ws = wb['라인별 추천태그']
    tags = {}
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=9, values_only=True):
        if not r[0]:
            continue
        line_id = str(r[0]).strip()
        tags[line_id] = dict(
            recipient=split_tags(r[3]), occasion=split_tags(r[4]), giftType=split_tags(r[5]),
            situation=str(r[6] or '').strip(), reason=str(r[7] or '').strip(),
            note=str(r[8] or '').strip() or None,
        )
    return tags


def main():
    tags = load_tags()
    images = scan_images()
    missing = [ln['id'] for ln in (LINES + EXTRA_LINES) if ln['id'] not in tags]
    if missing:
        print(f"!! gift-finder-tags.xlsx에 태그가 없는 라인 {len(missing)}개 (건너뜀): {', '.join(missing)}")

    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb['Sheet1']
    skus = []
    bad_urls = []
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=5, values_only=True):
        if not r[0]:
            continue
        url = str(r[4]).strip() if len(r) > 4 and r[4] else None
        if url and not URL_RE.match(url):
            bad_urls.append((r[0], url))
            url = None
        skus.append(dict(name=str(r[0]).strip(), price=r[1], mat=str(r[2] or '').strip(), url=url))

    buckets = {ln['id']: [] for ln in LINES}
    unmatched = []
    for s in skus:
        hit = next((ln for ln in LINES if matches(s['name'], ln['match'])), None)
        if hit:
            buckets[hit['id']].append(s)
        else:
            unmatched.append(s['name'])

    lines_out = []
    for ln in LINES:
        group = buckets[ln['id']]
        if not group:
            print(f"!! 빈 라인: {ln['name']}")
            continue
        t = tags.get(ln['id'])
        if not t:
            continue
        prices = [g['price'] for g in group]
        mats = sorted({g['mat'] for g in group})
        urls_in_line = sorted({g['url'] for g in group if g['url']})
        if len(urls_in_line) > 1:
            print(f"?? 라인 '{ln['name']}' 안에 서로 다른 링크가 {len(urls_in_line)}개 있음 — 첫 번째만 사용:")
            for u in urls_in_line:
                print(f"     {u}")
        line_url = urls_in_line[0] if urls_in_line else URL_MAP.get(ln['id'])
        variants = []
        for g in group:
            v = OrderedDict(sku=g['name'], price=g['price'])
            for key, pool in (('color', COLORS), ('size', SIZES), ('tier', TIERS), ('option', OPTS)):
                val = extract(g['name'], pool)
                if val:
                    v[key] = val
            variants.append(v)
        lines_out.append(OrderedDict([
            ("id", ln['id']), ("name", ln['name']), ("category", ln['cat']),
            ("material", mats[0] if len(mats) == 1 else mats),
            ("priceMin", min(prices)), ("priceMax", max(prices)),
            ("priceBands", sorted({price_band(p) for p in prices},
                                  key=lambda b: ["1만원대", "2만원대", "3만원대", "4만원대", "5만원 이상"].index(b))),
            ("recipient", t['recipient']), ("occasion", t['occasion']),
            ("giftType", t['giftType']),
            ("situation", t['situation']), ("reason", t['reason']),
            ("url", line_url),
            ("image", images.get(ln['id'])),
            ("skuCount", len(group)), ("variants", variants),
        ]))

    for ln in EXTRA_LINES:
        t = tags.get(ln['id'])
        if not t:
            continue
        prices = [v['price'] for v in ln['variants']]
        lines_out.append(OrderedDict([
            ("id", ln['id']), ("name", ln['name']), ("category", ln['cat']),
            ("material", "체험 프로그램"),
            ("priceMin", min(prices)), ("priceMax", max(prices)),
            ("priceBands", sorted({price_band(p) for p in prices},
                                  key=lambda b: ["1만원대", "2만원대", "3만원대", "4만원대", "5만원 이상"].index(b))),
            ("recipient", t['recipient']), ("occasion", t['occasion']),
            ("giftType", t['giftType']),
            ("situation", t['situation']), ("reason", t['reason']),
            ("url", URL_MAP.get(ln['id'])),
            ("image", images.get(ln['id'])),
            ("note", t['note']),
            ("skuCount", len(ln['variants'])), ("variants", ln['variants']),
        ]))

    recipients, occasions, gift_types = [], [], []
    for l in lines_out:
        for v in l['recipient']:
            if v not in recipients: recipients.append(v)
        for v in l['occasion']:
            if v not in occasions: occasions.append(v)
        for v in l['giftType']:
            if v not in gift_types: gift_types.append(v)

    doc = OrderedDict([
        ("meta", OrderedDict([
            ("updated", "2026-08-06"),
            ("source", "테차상품가격리스트.xlsx (가격) + gift-finder-tags.xlsx (추천 태그)"),
            ("skuCount", len(skus)),
            ("lineCount", len(lines_out)),
            ("note", "제품 라인 단위로 추천 태그를 관리한다. 색상·사이즈·전원방식 등 변형은 variants에 담긴다. "
                     "가격은 원(KRW). 태그를 고칠 때는 docs/gift-finder-tags.xlsx를 고치고 스크립트를 다시 돌린다 — "
                     "이 JSON을 직접 편집하지 말 것(다음 실행 때 덮어써진다)."),
        ])),
        ("axes", OrderedDict([
            ("recipient", recipients),
            ("occasion", occasions),
            ("giftType", gift_types),
            ("priceBand", ["1만원대", "2만원대", "3만원대", "4만원대", "5만원 이상"]),
        ])),
        ("lines", lines_out),
    ])

    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(doc, f, ensure_ascii=False, indent=2)

    covered = sum(l['skuCount'] for l in lines_out if l['id'] != 'flower-class')
    print(f"라인 {len(lines_out)}개 / SKU {covered}개 매칭 (원본 {len(skus)}개)")
    if bad_urls:
        print(f"\n!! 유효하지 않은 링크 {len(bad_urls)}개 (http로 시작 안 함 — 무시하고 진행):")
        for name, url in bad_urls:
            print(f"   - {name}: '{url}'")
    no_url = [l['name'] for l in lines_out if not l.get('url')]
    if no_url:
        print(f"\n-- 아직 링크 없는 라인 {len(no_url)}개: {', '.join(no_url)}")

    no_img = [l['id'] for l in lines_out if not l.get('image')]
    have = len(lines_out) - len(no_img)
    print(f"\n-- 상품 사진 {have}/{len(lines_out)}개 연결됨 (assets/products/)")
    if no_img:
        print(f"   아직 없는 파일: {', '.join(n + '.jpg' for n in no_img)}")
    orphan = sorted(set(images) - {l['id'] for l in lines_out})
    if orphan:
        print(f"\n!! assets/products/ 안에 라인 id와 안 맞는 파일 {len(orphan)}개 (오타 확인):")
        for o in orphan:
            print(f"   - {images[o]}")
    if unmatched:
        print(f"\n!! 미분류 {len(unmatched)}개:")
        for u in unmatched:
            print("   -", u)
    else:
        print("미분류 없음 — 147개 전부 매칭됨")
    print(f"\n저장: {OUT}")
    for l in lines_out:
        print(f"  {l['skuCount']:3d}종 {l['priceMin']:>6,}~{l['priceMax']:>6,}원  {l['name']}")


if __name__ == '__main__':
    main()
